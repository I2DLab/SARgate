"""
=============================
lmm_file_reader.py
=============================

Robust readers for CSV/TSV/TXT and XLSX files.
- Autodetect separator with the 'python' engine.
- Try multiple encodings.
- Skip bad lines.
- For XLSX, inspect worksheets with openpyxl and extract the most likely table.
"""

# =============================================================================
# =============================================================================
# 1. Import module dependencies
# 2. Read text robust
# 3. Read excel robust

import os
import re
import sys
import warnings
import zipfile
import pandas as pd
pd.set_option("future.no_silent_downcasting", True)
from typing import Any
from xml.etree import ElementTree as ET


# -----------------------------------------------------------------------------
# 2. Read text robust
# -----------------------------------------------------------------------------
def _read_text_robust(path: str) -> Any:
    encodings = ("utf-8-sig", "cp1252", "latin-1")
    last_exc = None
    for enc in encodings:
        try:
            return pd.read_csv(path, sep=None, engine="python",
                                encoding=enc, on_bad_lines="skip")
        except Exception as e:
            last_exc = e
    raise last_exc


# -----------------------------------------------------------------------------
# 3. Read excel robust
# -----------------------------------------------------------------------------
def _read_excel_robust(path: str) -> Any:

    ext = os.path.splitext(path)[1].lower()
    if ext != ".xlsx":
        raise ValueError(f"Unsupported Excel format '{ext}'. Only .xlsx files are supported.")

    base = os.path.basename(path)
    if base.startswith("~$"):
        raise ValueError("Selected a temporary Excel lock file (~$...). Choose the real .xlsx workbook.")

    def _normalize_cell(value: Any) -> Any:
        """
        Normalize a worksheet cell value for dataframe construction.
        """
        if value is None:
            return ""
        return value

    def _largest_bbox(ws: Any) -> Any:
        min_r = min_c = 10**9
        max_r = max_c = 0
        any_data = False
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c_idx, v in enumerate(row, start=1):
                if v is not None and str(v).strip() != "":
                    any_data = True
                    if r_idx < min_r:
                        min_r = r_idx
                    if r_idx > max_r:
                        max_r = r_idx
                    if c_idx < min_c:
                        min_c = c_idx
                    if c_idx > max_c:
                        max_c = c_idx
        if not any_data:
            return None
        return (min_r, min_c, max_r, max_c)

    def _trim_block(block: list[list[Any]]) -> list[list[Any]]:
        """
        Remove empty outer rows and columns from a 2D block.
        """
        if not block:
            return []

        non_empty_rows = [
            idx for idx, row in enumerate(block)
            if any(str(value).strip() != "" for value in row)
        ]
        if not non_empty_rows:
            return []

        non_empty_cols = [
            idx for idx in range(len(block[0]))
            if any(str(row[idx]).strip() != "" for row in block)
        ]
        if not non_empty_cols:
            return []

        first_row, last_row = non_empty_rows[0], non_empty_rows[-1]
        first_col, last_col = non_empty_cols[0], non_empty_cols[-1]
        return [
            row[first_col:last_col + 1]
            for row in block[first_row:last_row + 1]
        ]

    def _score_header_row(row: list[Any]) -> float:
        """
        Score how likely a row is to contain column headers.
        """
        if not row:
            return 0.0

        cleaned = [str(value).strip() for value in row]
        non_empty = [value for value in cleaned if value]
        if not non_empty:
            return 0.0

        text_like = sum(
            1 for value in non_empty
            if any(ch.isalpha() for ch in value) and not value.replace(".", "", 1).isdigit()
        )
        unique_ratio = len(set(non_empty)) / len(non_empty)
        filled_ratio = len(non_empty) / len(cleaned)
        return (text_like / len(non_empty)) + unique_ratio + filled_ratio

    def _make_unique_headers(header_row: list[Any], width: int) -> list[str]:
        """
        Convert the chosen header row into unique dataframe column names.
        """
        headers: list[str] = []
        counts: dict[str, int] = {}
        for idx in range(width):
            raw = str(header_row[idx]).strip() if idx < len(header_row) else ""
            base = raw if raw else f"col_{idx + 1}"
            count = counts.get(base, 0)
            counts[base] = count + 1
            headers.append(base if count == 0 else f"{base}_{count + 1}")
        return headers

    def _block_to_df(block: list[list[Any]]) -> Any:
        """
        Convert a trimmed 2D block into a dataframe.
        """
        trimmed = _trim_block(block)
        if not trimmed:
            return None

        width = max(len(row) for row in trimmed)
        padded = [row + [""] * (width - len(row)) for row in trimmed]

        candidate_rows = padded[: min(5, len(padded))]
        header_idx = max(
            range(len(candidate_rows)),
            key=lambda idx: _score_header_row(candidate_rows[idx]),
            default=0,
        )

        columns = _make_unique_headers(padded[header_idx], width)
        data = padded[header_idx + 1:] if header_idx < len(padded) - 1 else []
        if not data:
            data = padded
            columns = [f"col_{idx + 1}" for idx in range(width)]

        df_local = pd.DataFrame(data, columns=columns)
        df_local = df_local.replace("", pd.NA).dropna(axis=1, how="all").dropna(how="all")
        return df_local if not df_local.empty else None

    def _bbox_to_df(ws: Any, bbox: Any) -> Any:
        (min_r, min_c, max_r, max_c) = bbox
        block = []
        for r in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True):
            block.append([_normalize_cell(v) for v in r])
        return _block_to_df(block)

    def _load_workbook_quietly(data_only: bool) -> Any:
        """
        Load an XLSX workbook while suppressing known non-fatal openpyxl warnings.
        """
        from openpyxl import load_workbook

        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=r"File contains an invalid specification for \d+\. This will be removed",
                category=UserWarning,
            )
            return load_workbook(path, data_only=data_only, read_only=True)

    def _should_try_openpyxl() -> bool:
        """
        Decide whether the optional openpyxl reader should be used.

        Python 3.13 Linux environments can segfault while importing openpyxl
        before Python can raise an exception, so the XML reader is preferred
        there unless the user explicitly opts in.
        """
        override = os.environ.get("SARGATE_USE_OPENPYXL", "").strip().lower()
        if override in {"1", "true", "yes", "on"}:
            return True
        if override in {"0", "false", "no", "off"}:
            return False
        return sys.version_info < (3, 13)

    def _xml_local_name(tag: str) -> str:
        """
        Extract the local XML name without namespace.
        """
        return tag.rsplit("}", 1)[-1]

    def _excel_col_to_index(cell_ref: str) -> int:
        """
        Convert an Excel cell reference like 'BC12' to a 1-based column index.
        """
        letters = "".join(ch for ch in cell_ref if ch.isalpha()).upper()
        idx = 0
        for ch in letters:
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx

    def _load_shared_strings(zf: Any) -> list[str]:
        """
        Load the shared string table used by many XLSX files.
        """
        try:
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        except KeyError:
            return []

        shared_strings: list[str] = []
        for si in root:
            if _xml_local_name(si.tag) != "si":
                continue
            parts: list[str] = []
            for node in si.iter():
                if _xml_local_name(node.tag) == "t":
                    parts.append(node.text or "")
            shared_strings.append("".join(parts))
        return shared_strings

    def _load_workbook_sheet_targets(zf: Any) -> list[tuple[str, str]]:
        """
        Read workbook metadata and resolve worksheet XML targets.
        """
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

        rel_map: dict[str, str] = {}
        for rel in rels_root:
            if _xml_local_name(rel.tag) != "Relationship":
                continue
            rel_id = rel.attrib.get("Id", "")
            target = rel.attrib.get("Target", "")
            if rel_id and target:
                rel_map[rel_id] = f"xl/{target.lstrip('/')}"

        sheets: list[tuple[str, str]] = []
        for node in workbook_root.iter():
            if _xml_local_name(node.tag) != "sheet":
                continue
            rel_id = ""
            for key, value in node.attrib.items():
                if key.endswith("}id") or key == "id":
                    rel_id = value
                    break
            name = node.attrib.get("name", "Sheet")
            target = rel_map.get(rel_id)
            if target:
                sheets.append((name, target))
        return sheets

    def _parse_xml_sheet(zf: Any, target: str, shared_strings: list[str]) -> Any:
        """
        Parse a worksheet XML file into a dataframe.
        """
        root = ET.fromstring(zf.read(target))
        rows_by_idx: dict[int, dict[int, Any]] = {}

        for row_node in root.iter():
            if _xml_local_name(row_node.tag) != "row":
                continue

            row_idx = int(row_node.attrib.get("r", len(rows_by_idx) + 1))
            row_values: dict[int, Any] = {}

            for cell in row_node:
                if _xml_local_name(cell.tag) != "c":
                    continue
                cell_ref = cell.attrib.get("r", "")
                col_idx = _excel_col_to_index(cell_ref) if cell_ref else len(row_values) + 1
                cell_type = cell.attrib.get("t", "")

                value_text = None
                inline_text_parts: list[str] = []
                for child in cell.iter():
                    child_name = _xml_local_name(child.tag)
                    if child_name == "v" and value_text is None:
                        value_text = child.text
                    elif child_name == "t":
                        inline_text_parts.append(child.text or "")

                value: Any = value_text
                if cell_type == "s":
                    try:
                        value = shared_strings[int(value_text)] if value_text is not None else ""
                    except Exception:
                        value = value_text or ""
                elif cell_type == "inlineStr":
                    value = "".join(inline_text_parts)
                elif cell_type == "b":
                    value = value_text == "1"
                elif cell_type in ("str", "e"):
                    value = value_text or ""
                elif value_text is None and inline_text_parts:
                    value = "".join(inline_text_parts)
                elif value_text is None:
                    value = ""
                else:
                    try:
                        numeric = float(value_text)
                        value = int(numeric) if numeric.is_integer() else numeric
                    except Exception:
                        value = value_text

                row_values[col_idx] = value

            if row_values:
                rows_by_idx[row_idx] = row_values

        if not rows_by_idx:
            return None

        min_row = min(rows_by_idx)
        max_row = max(rows_by_idx)
        max_col = max(max(cols) for cols in rows_by_idx.values())
        block: list[list[Any]] = []
        for row_idx in range(min_row, max_row + 1):
            row_map = rows_by_idx.get(row_idx, {})
            block.append([row_map.get(col_idx, "") for col_idx in range(1, max_col + 1)])

        return _block_to_df(block)

    best_df = None
    best_score = (-1, -1)
    last_err = "no readable worksheet found"

    try:
        with zipfile.ZipFile(path, "r") as zf:
            shared_strings = _load_shared_strings(zf)
            for _, target in _load_workbook_sheet_targets(zf):
                try:
                    df = _parse_xml_sheet(zf, target, shared_strings)
                except Exception:
                    continue
                if df is None:
                    continue
                score = (df.shape[0] * df.shape[1], df.shape[1])
                if score > best_score:
                    best_df = df
                    best_score = score
    except Exception as exc:
        last_err = str(exc)

    if best_df is not None:
        return best_df.reset_index(drop=True)

    if _should_try_openpyxl():
        for data_only in (True, False):
            try:
                wb = _load_workbook_quietly(data_only=data_only)
                for ws in getattr(wb, "worksheets", []):
                    bbox = _largest_bbox(ws)
                    if not bbox:
                        continue
                    df = _bbox_to_df(ws, bbox)
                    if df is None:
                        continue
                    score = (df.shape[0] * df.shape[1], df.shape[1])
                    if score > best_score:
                        best_df = df
                        best_score = score
            except Exception as exc:
                last_err = str(exc)

            if best_df is not None:
                return best_df.reset_index(drop=True)
    else:
        last_err = "openpyxl skipped for this Python runtime"

    raise ValueError(
        "Workbook contains no tabular data detectable in any worksheet. "
        "Open the file in Excel or LibreOffice, copy the visible table into a new sheet using Paste Values, then save again as .xlsx or export as CSV. "
        f"[openpyxl status: {last_err}]"
    )
